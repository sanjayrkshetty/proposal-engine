#!/usr/bin/env python3
"""Pipeline Report — scans outputs/ and generates Markdown or Excel report."""
import sys
import os
import json
import re
from pathlib import Path
from datetime import datetime

import click
from rich.console import Console
from rich.table import Table

console = Console()

# Excel dark theme colours
EXCEL_BG = "0D1117"
EXCEL_HEADER_BG = "1C2333"
EXCEL_TEXT = "C9D1D9"
EXCEL_ACCENT = "58A6FF"

OUTPUTS_DIR = Path("outputs")


def scan_outputs(bu_filter: str = None, region_filter: str = None) -> list:
    """Scan outputs/ directory and extract proposal metadata."""
    proposals = []
    if not OUTPUTS_DIR.exists():
        return proposals

    for run_dir in sorted(OUTPUTS_DIR.iterdir()):
        if not run_dir.is_dir():
            continue

        # Parse directory name: YYYY-MM-DD_ClientName_BU_ServiceCode
        parts = run_dir.name.split("_")
        if len(parts) < 3:
            continue

        date_str = parts[0]
        bu_from_dir = parts[2] if len(parts) > 2 else "UNKNOWN"
        service_from_dir = parts[3] if len(parts) > 3 else "UNKNOWN"

        # Apply filters
        if bu_filter and bu_from_dir.upper() != bu_filter.upper():
            continue

        # Load metadata from proposal_meta.json if exists
        meta_file = run_dir / "proposal_meta.json"
        pricing_file = run_dir / "pricing_table.json"

        proposal = {
            "run_id": run_dir.name,
            "date": date_str,
            "client": "Unknown",
            "bu": bu_from_dir,
            "service": service_from_dir,
            "proposal_number": "—",
            "region": "Unknown",
            "tier": "—",
            "acv_inr": 0,
            "acv_display": "—",
            "currency": "INR",
        }

        if meta_file.exists():
            try:
                meta = json.loads(meta_file.read_text(encoding="utf-8"))
                proposal["client"] = meta.get("client") or proposal["client"]
                proposal["bu"] = meta.get("service", service_from_dir).split("-")[0] if meta.get("service") else bu_from_dir
                proposal["service"] = meta.get("service") or service_from_dir
                proposal["proposal_number"] = meta.get("proposal_number") or "—"
                proposal["region"] = meta.get("region") or "Unknown"
                proposal["tier"] = meta.get("tier") or "—"
                proposal["bu"] = _infer_bu(proposal["service"])
            except (json.JSONDecodeError, Exception):
                pass

        if pricing_file.exists():
            try:
                pricing = json.loads(pricing_file.read_text(encoding="utf-8"))
                acv = pricing.get("total_inr", 0)
                currency = pricing.get("currency", "INR")
                total_display = pricing.get("total_display", 0)
                sym = {"INR": "₹", "USD": "$", "SGD": "S$"}.get(currency, "")
                proposal["acv_inr"] = acv
                proposal["acv_display"] = f"{sym}{total_display:,.0f} {currency}"
                proposal["currency"] = currency
            except (json.JSONDecodeError, Exception):
                pass

        # Apply region filter
        if region_filter and proposal["region"].upper() != region_filter.upper():
            continue

        proposals.append(proposal)

    return proposals


def _infer_bu(service_code: str) -> str:
    if not service_code:
        return "UNKNOWN"
    s = service_code.upper()
    if s.startswith("DFIR") or s in ("IFI", "CA", "BAS"):
        return "DFIR"
    elif s.startswith("MXDR"):
        return "MXDR"
    elif s.startswith("DPG"):
        return "DPG"
    elif s.startswith("CTS"):
        return "CTS"
    elif s.startswith("INST"):
        return "INSTITUTE"
    elif s.startswith("QTM"):
        return "QUANTUM"
    return "UNKNOWN"


def render_markdown(proposals: list):
    """Render pipeline report as Rich Markdown table to console."""
    if not proposals:
        console.print("[yellow]No proposals found in outputs/ directory.[/yellow]")
        return

    table = Table(title="SISA Proposal Pipeline Report", show_lines=True)
    table.add_column("Date", style="dim")
    table.add_column("Client", style="bold cyan")
    table.add_column("BU", style="blue")
    table.add_column("Service", style="yellow")
    table.add_column("Proposal No.", style="green")
    table.add_column("Region", style="magenta")
    table.add_column("Tier", style="white")
    table.add_column("ACV", style="bold green")

    for p in proposals:
        table.add_row(
            p["date"],
            p["client"],
            p["bu"],
            p["service"],
            p["proposal_number"],
            p["region"],
            p["tier"],
            p["acv_display"],
        )

    console.print(table)
    console.print(f"\n[dim]Total proposals: {len(proposals)}[/dim]")

    # Also print raw markdown table
    console.print("\n[bold]Markdown Table:[/bold]")
    headers = ["Date", "Client", "BU", "Service", "Proposal No.", "Region", "Tier", "ACV"]
    header_line = "| " + " | ".join(headers) + " |"
    sep_line = "|" + "|".join(["---" for _ in headers]) + "|"
    print(header_line)
    print(sep_line)
    for p in proposals:
        row = [p["date"], p["client"], p["bu"], p["service"],
               p["proposal_number"], p["region"], p["tier"], p["acv_display"]]
        print("| " + " | ".join(row) + " |")


def render_excel(proposals: list, output_path: str = "reports/pipeline_report.xlsx"):
    """Render pipeline report as Excel file with dark theme."""
    try:
        import openpyxl
        from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
    except ImportError:
        console.print("[red]openpyxl not installed. Run: pip install openpyxl[/red]")
        return

    wb = openpyxl.Workbook()
    wb.remove(wb.active)  # Remove default sheet

    # Style helpers
    def make_header_font():
        return Font(bold=True, color=EXCEL_ACCENT, size=11)

    def make_body_font():
        return Font(color=EXCEL_TEXT, size=10)

    def make_header_fill():
        return PatternFill(start_color=EXCEL_HEADER_BG, end_color=EXCEL_HEADER_BG, fill_type="solid")

    def make_body_fill():
        return PatternFill(start_color=EXCEL_BG, end_color=EXCEL_BG, fill_type="solid")

    def make_border():
        side = Side(style="thin", color="30363D")
        return Border(left=side, right=side, top=side, bottom=side)

    COLUMNS = ["Date", "Client", "BU", "Service", "Proposal No.", "Region", "Tier", "ACV"]

    def write_sheet(ws, data, title):
        # Tab colour
        ws.sheet_properties.tabColor = EXCEL_ACCENT

        # Title row
        ws.append([title])
        title_cell = ws.cell(1, 1)
        title_cell.font = Font(bold=True, color=EXCEL_ACCENT, size=14)
        title_cell.fill = make_body_fill()
        ws.merge_cells(f"A1:{get_column_letter(len(COLUMNS))}1")
        ws.row_dimensions[1].height = 28

        # Header row
        ws.append(COLUMNS)
        for col_num, _ in enumerate(COLUMNS, 1):
            cell = ws.cell(2, col_num)
            cell.font = make_header_font()
            cell.fill = make_header_fill()
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            cell.border = make_border()
        ws.row_dimensions[2].height = 22

        # Data rows
        for row_idx, p in enumerate(data, 3):
            row_data = [p["date"], p["client"], p["bu"], p["service"],
                        p["proposal_number"], p["region"], p["tier"], p["acv_display"]]
            ws.append(row_data)
            for col_num in range(1, len(COLUMNS) + 1):
                cell = ws.cell(row_idx, col_num)
                cell.font = make_body_font()
                cell.fill = make_body_fill()
                cell.alignment = Alignment(vertical="center")
                cell.border = make_border()
            ws.row_dimensions[row_idx].height = 18

        # Column widths
        col_widths = [12, 28, 10, 20, 20, 14, 12, 18]
        for col_num, width in enumerate(col_widths, 1):
            ws.column_dimensions[get_column_letter(col_num)].width = width

    # Summary sheet
    ws_summary = wb.create_sheet("Summary")
    write_sheet(ws_summary, proposals, "SISA Proposal Pipeline — All BUs")

    # Per-BU sheets
    bu_groups: dict = {}
    for p in proposals:
        bu = p["bu"]
        bu_groups.setdefault(bu, []).append(p)

    for bu, bu_proposals in sorted(bu_groups.items()):
        ws_bu = wb.create_sheet(bu[:30])  # Sheet name max 31 chars
        write_sheet(ws_bu, bu_proposals, f"SISA Pipeline — {bu}")

    # Ensure reports/ directory exists
    Path(output_path).parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)
    console.print(f"[green]✓ Excel report saved:[/green] {output_path}")


@click.command()
@click.option("--format", "fmt", default="markdown",
              type=click.Choice(["markdown", "excel"], case_sensitive=False),
              help="Output format")
@click.option("--bu", default=None, help="Filter by BU (DFIR, MXDR, DPG, CTS, INSTITUTE, QUANTUM)")
@click.option("--region", default=None, help="Filter by region")
@click.option("--output", "output_path", default="reports/pipeline_report.xlsx",
              help="Output path for Excel report")
def main(fmt, bu, region, output_path):
    """SISA Pipeline Report — scan outputs and generate report."""
    console.print(f"[bold cyan]SISA Pipeline Report[/bold cyan] — format: {fmt}")
    if bu:
        console.print(f"  Filter: BU = {bu}")
    if region:
        console.print(f"  Filter: Region = {region}")

    proposals = scan_outputs(bu_filter=bu, region_filter=region)

    if fmt == "markdown":
        render_markdown(proposals)
    elif fmt == "excel":
        render_excel(proposals, output_path)


if __name__ == "__main__":
    # Ensure we run from repo root
    repo_root = Path(__file__).parent.parent
    os.chdir(repo_root)
    main()
